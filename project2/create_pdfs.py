# Process deviations:
# - No price
# - No delivery date
# - Articlenumber didn't match
# - Wrong quantity
# - Other payment terms
# - Other delivery terms
# - Order number doesn't match

from faker import Faker
import faker_commerce

fake = Faker()
fake.add_provider(faker_commerce.Provider)
from datetime import datetime, date
import pandas as pd
from random import randint
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
from win32com import client
import copy

file_path = os.path.dirname(__file__)
os.chdir(file_path)


class FakePurchaseConfirmationCreator:
    def __init__(self, template_xlsx_path, logo_path):
        self.template_xlsx_path = template_xlsx_path
        self.logo_path = logo_path
        self.fake_decimails = ["00", "25", "50", "75", "95", "99"]
        self.fake_payment_terms_list = [14, 30, 60, 90]
        self.fake_delivery_terms_list = [
            "EXW - Ex Works",
            "FCA - Free Carrier",
            "FAS - Free Alongside Ship",
            "FOB - Free on Board",
            "CFR - Cost and Freight",
            "CIF - Cost, Insurance, and Freight",
            "CPT - Carriage Paid To",
            "CIP - Carriage and Insurance Paid To",
            "DAT - Delivered at Terminal",
            "DAP - Delivered at Place",
            "DDP - Delivered Duty Paid",
        ]
        self.purchaseheader_dict = dict(
            zip(
                ["ordernumber", "order_date", "delivery_terms", "payment_terms"],
                [[] for _ in range(4)],
            )
        )
        self.purchaserow_dict = dict(
            zip(
                [
                    "ordernumber",
                    "orderline",
                    "item",
                    "delivery_date",
                    "quantity",
                    "price_per_piece",
                ],
                [[] for _ in range(6)],
            )
        )

    def create_general_information(self, pdfnr):
        self.fake_doc_date = fake.date_between(start_date=date(2024, 2, 24), end_date=date(2024, 2, 28))
        self.fake_doc_date = self.fake_doc_date.strftime("%d-%m-%Y")
        self.fake_ordernr = "2410" + str(randint(100, 999))
        index_delivery_terms = randint(0, 10)
        self.fake_delivery_terms = self.fake_delivery_terms_list[index_delivery_terms]
        self.wrong_delivery_terms = index_delivery_terms - 1 if index_delivery_terms == 10 else index_delivery_terms + 1
        index_payment_terms = randint(0, 3)
        index_wrong_payment_terms = index_payment_terms - 1 if index_payment_terms == 3 else index_payment_terms + 1
        self.fake_payment_terms = f"Payment within {self.fake_payment_terms_list[index_payment_terms]} days."
        self.wrong_payment_terms = f"Payment within {self.fake_payment_terms_list[index_wrong_payment_terms]} days."
        self.fake_orderref = f"924{str(pdfnr).zfill(3)}" + str(randint(10, 99))
        self.wrong_orderref = "923" + self.fake_orderref[3:]
        self.fake_contactperson = fake.name()
        return (
            self.fake_doc_date,
            self.fake_ordernr,
            self.fake_delivery_terms,
            self.wrong_delivery_terms,
            self.fake_payment_terms,
            self.wrong_payment_terms,
            self.fake_orderref,
            self.wrong_orderref,
            self.fake_contactperson,
        )

    def create_orderlines(self):
        self.fake_articlenr = "123." + str(randint(100, 999)) + "." + str(randint(1000, 9999))
        self.wrong_articlenr = "123." + str(randint(100, 999)) + "." + str(randint(1000, 9999))
        self.fake_productname = fake.ecommerce_name()
        self.fake_item = "{} - {}".format(self.fake_articlenr, self.fake_productname)
        self.wrong_item = "{} - {}".format(self.wrong_articlenr, self.fake_productname)
        self.fake_delivery_date = fake.date_between(start_date=date(2024, 3, 1), end_date=date(2024, 6, 10))
        self.fake_delivery_date = self.fake_delivery_date.strftime("%d-%m-%Y")
        self.wrong_delivery_data = " "
        self.fake_qty = randint(1, 15)
        self.wrong_qty = self.fake_qty + 2
        self.fake_price_per_piece = float(str(randint(1, 999)) + "." + str(self.fake_decimails[randint(0, 5)]))
        self.wrong_price_per_piece = float(str(randint(1, 999)) + "." + str(self.fake_decimails[randint(0, 5)]))
        return (
            self.fake_item,
            self.fake_delivery_date,
            self.fake_qty,
            self.fake_price_per_piece,
            self.wrong_price_per_piece,
            self.wrong_delivery_data,
            self.wrong_qty,
            self.wrong_item,
        )

    def determine_orderlines(self, fake_orderline, pdfnr, error):
        self.create_orderlines()
        self.orderline_data_correct = [
            self.fake_orderref,
            fake_orderline,
            self.fake_item,
            self.fake_delivery_date,
            self.fake_qty,
            self.fake_price_per_piece,
            self.fake_delivery_terms,
            self.fake_payment_terms,
        ]
        if (pdfnr > 40) & (pdfnr <= 50) & (error == True):
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.wrong_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]
        elif (pdfnr > 50) & (pdfnr <= 75) & (error == True):
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.wrong_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]
        elif (pdfnr > 75) & (pdfnr <= 85) & (error == True):
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.wrong_delivery_data,
                self.fake_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]
        elif (pdfnr > 85) & (pdfnr <= 100) & (error == True):
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.wrong_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]
        elif (pdfnr > 100) & (pdfnr <= 110):
            self.orderline_data_list = [
                self.wrong_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]
        elif (pdfnr > 110) & (pdfnr <= 125):
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.fake_price_per_piece,
                self.wrong_delivery_terms,
                self.fake_payment_terms,
            ]
        elif pdfnr > 125:
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.wrong_payment_terms,
            ]

        else:
            self.orderline_data_list = [
                self.fake_orderref,
                fake_orderline,
                self.fake_item,
                self.fake_delivery_date,
                self.fake_qty,
                self.fake_price_per_piece,
                self.fake_delivery_terms,
                self.fake_payment_terms,
            ]

        return self.orderline_data_correct, self.orderline_data_list

    def create_xlsx_workbook(self, pdfnr):
        self.wb = load_workbook(self.template_xlsx_path)
        self.ws = self.wb["PurchaseOrder"]
        logo = Image(self.logo_path)
        self.ws.add_image(logo, "A1")

        for fake_orderline in range(1, randint(2, 10)):
            error = [True, False][randint(0, 1)]
            fake_orderline_adjusted = "{}.".format(fake_orderline)
            orderline_data_list_correct, orderline_data_list = self.determine_orderlines(
                fake_orderline_adjusted, pdfnr, error
            )
            self.purchaserow_dict["ordernumber"].append(orderline_data_list_correct[0])
            self.purchaserow_dict["orderline"].append(orderline_data_list_correct[1])
            self.purchaserow_dict["item"].append(orderline_data_list_correct[2])
            self.purchaserow_dict["delivery_date"].append(orderline_data_list_correct[3])
            self.purchaserow_dict["quantity"].append(orderline_data_list_correct[4])
            self.purchaserow_dict["price_per_piece"].append(orderline_data_list_correct[5])

            for idx, data in enumerate(orderline_data_list[1:6]):
                self.ws.cell(fake_orderline + 7, idx + 1).value = data

        self.purchaseheader_dict["ordernumber"].append(orderline_data_list_correct[0])
        self.purchaseheader_dict["order_date"].append(self.fake_doc_date)
        self.purchaseheader_dict["delivery_terms"].append(orderline_data_list_correct[6])
        self.purchaseheader_dict["payment_terms"].append(orderline_data_list_correct[7])

        self.ws["F3"] = self.fake_doc_date
        self.ws["F4"] = self.fake_ordernr
        self.ws["F5"] = orderline_data_list[0]
        self.ws["B24"] = orderline_data_list[6]
        self.ws["B25"] = orderline_data_list[7]
        self.ws["B26"] = self.fake_contactperson

        dims = {}
        for row in self.ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.ws.column_dimensions[col].width = value + 2
        self.confirmation_xlsx_path = "{}/fake_confirmations/xlsx/Confirmation_{}_{}.xlsx".format(
            file_path, self.fake_ordernr, self.fake_orderref
        )
        self.wb.save(self.confirmation_xlsx_path)
        self.wb.close()
        return (
            self.confirmation_xlsx_path,
            self.purchaseheader_dict,
            self.purchaserow_dict,
        )

    def create_pdf_confirmation(self):
        excel = client.Dispatch("Excel.Application")
        sheets = excel.Workbooks.Open(self.confirmation_xlsx_path)
        work_sheets = sheets.Worksheets[0]
        work_sheets.PageSetup.Zoom = False
        work_sheets.PageSetup.FitToPagesWide = 1
        work_sheets.PageSetup.FitToPagesTall = 1
        work_sheets.PageSetup.PrintArea = work_sheets.UsedRange.Address

        pdf_path = "{}/fake_confirmations/pdfs/Confirmation_{}_{}.pdf".format(
            file_path, self.fake_ordernr, self.fake_orderref
        )
        work_sheets.ExportAsFixedFormat(0, pdf_path)

    def return_dictionaries(self):
        self.purchaseheader_df, self.purchaserow_df = pd.DataFrame(self.purchaseheader_dict), pd.DataFrame(
            self.purchaserow_dict
        )
        return (
            self.purchaseheader_df,
            self.purchaserow_df,
        )


def main(pdfnr):
    creator = FakePurchaseConfirmationCreator(
        template_xlsx_path="{}/Template/xlsx/purchase-order.xlsx".format(file_path),
        logo_path="{}/Template/png/logo.png".format(file_path),
    )
    creator.create_general_information(pdfnr)
    creator.create_xlsx_workbook(pdfnr)
    creator.create_pdf_confirmation()
    purchaseheader_temp_df, purchaserow_temp_df = creator.return_dictionaries()
    return purchaseheader_temp_df, purchaserow_temp_df


if __name__ == "__main__":
    total_purchaseheader_df, total_purchaserow_df = pd.DataFrame(), pd.DataFrame()
    for pdfnr in range(175):
        purchaseheader_temp_df, purchaserow_temp_df = main(pdfnr)
        total_purchaseheader_df = pd.concat([total_purchaseheader_df, purchaseheader_temp_df])
        total_purchaserow_df = pd.concat([total_purchaserow_df, purchaserow_temp_df])
    total_purchaseheader_df.to_parquet(
        f"{file_path}/fake_confirmations/snappy/total_purchaseheader_df.snappy.parquet",
        engine="pyarrow",
        compression="snappy",
    )
    total_purchaserow_df.to_parquet(
        f"{file_path}/fake_confirmations/snappy/total_purchaserow_df.snappy.parquet",
        engine="pyarrow",
        compression="snappy",
    )
os.system("taskkill /f /im excel.exe")
